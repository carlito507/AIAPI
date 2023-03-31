from fastapi import UploadFile, File, APIRouter, HTTPException, status
from chardet.universaldetector import UniversalDetector
from PyPDF2 import PdfFileReader
import openpyxl
import xlrd
from io import BytesIO
from app.core.config import ALLOWED_EXTENSIONS

import logging

logger = logging.getLogger(__name__)

# Create a router
loaders_router = APIRouter()


def load_text_file_from_disk(file_path: str) -> str:
    with open(file_path, 'r') as file:
        return file.read()


def load_text_file(file: UploadFile) -> str:
    detector = UniversalDetector()
    for line in file.file:
        detector.feed(line)
        if detector.done:
            break
    detector.close()
    encoding = detector.result['encoding']
    file.file.seek(0)
    content = file.file.read().decode(encoding)
    return content


def load_pdf_file_from_disk(file_path: str) -> str:
    with open(file_path, 'rb') as file:
        pdf_reader = PdfFileReader(file)
        content = ""
        for i in range(pdf_reader.getNumPages()):
            content += pdf_reader.getPage(i).extractText()
        return content


def load_pdf_file(file: UploadFile) -> str:
    pdf_reader = PdfFileReader(BytesIO(file.file.read()))
    content = ""
    for i in range(pdf_reader.getNumPages()):
        content += pdf_reader.getPage(i).extractText()
    return content


def load_excel_file_from_disk(file_path: str) -> list:
    if file_path.endswith('.xlsx'):
        wb = openpyxl.load_workbook(file_path)
    elif file_path.endswith('.xls'):
        wb = xlrd.open_workbook(file_path)
    else:
        raise ValueError("Invalid file format. Only .xlsx and .xls files are supported.")
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data


def load_excel_file(file: UploadFile) -> list:
    if file.filename.endswith('.xlsx'):
        wb = openpyxl.load_workbook(BytesIO(file.file.read()))
    elif file.filename.endswith('.xls'):
        wb = xlrd.open_workbook(BytesIO(file.file.read()))
    else:
        raise ValueError("Invalid file format. Only .xlsx and .xls files are supported.")
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)
    return data


@loaders_router.post("/upload_file/")
async def upload_file(file: UploadFile = File(...)):
    file_ext = file.filename.split('.')[-1]
    if file.content_type == "text/plain" or f".{file_ext}" in ALLOWED_EXTENSIONS:
        content = load_text_file(file)
    elif file.content_type == "application/pdf":
        content = load_pdf_file(file)
    elif file.content_type in ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "application/vnd.ms-excel"]:
        content = load_excel_file(file)
    else:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid file format. Only .txt, .pdf, .xlsx and .xls files are supported.")
    return {"filename": file.filename, "content": content}